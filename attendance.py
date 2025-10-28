import cv2
import pickle
from datetime import datetime
from ultralytics import YOLO
from deepface import DeepFace
from openpyxl import Workbook, load_workbook

# Load classifier + labels
with open("face_classifier.pkl", "rb") as f:
    clf, label_ids = pickle.load(f)

id_to_name = {v: k for k, v in label_ids.items()}

# Load YOLOv8 detector
model = YOLO("yolov8n.pt")

# Excel path
excel_path = "attendance.xlsx"

# Initialize Excel if not exists
try:
    wb = load_workbook(excel_path)
    ws = wb.active
except:
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Status", "Emotion", "Date"])
    wb.save(excel_path)

# Attendance tracker
today = datetime.now().strftime("%Y-%m-%d")
attendance = {name: ["Absent", "-", today] for name in id_to_name.values()}

# Recognition function
def recognize_face(face_img):
    try:
        embedding = DeepFace.represent(
            face_img, model_name="ArcFace", enforce_detection=False
        )[0]["embedding"]
        pred = clf.predict([embedding])[0]
        return id_to_name[pred]
    except:
        return None

# Open webcam
cap = cv2.VideoCapture(0)

print("📷 Starting YOLOv8 + SVM Attendance System... Press 'q' to quit")

while True:
    ret, frame = cap.read()
    if not ret:
        break

    results = model(frame, verbose=False)

    for box in results[0].boxes.xyxy:
        x1, y1, x2, y2 = map(int, box[:4])
        face_img = frame[y1:y2, x1:x2]

        name = recognize_face(face_img)
        if name:
            try:
                # Emotion detection
                analysis = DeepFace.analyze(
                    face_img, actions=['emotion'], enforce_detection=False
                )
                emotion = analysis[0]['dominant_emotion']
            except:
                emotion = "-"

            attendance[name] = ["Present", emotion, today]
            cv2.putText(frame, f"{name} ({emotion})", (x1, y1-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,255,0), 2)
        else:
            cv2.putText(frame, "Unknown", (x1, y1-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,0,255), 2)

        cv2.rectangle(frame, (x1,y1), (x2,y2), (255,0,0), 2)

    cv2.imshow("YOLOv8 + SVM Attendance System", frame)
    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()

# Save attendance to Excel
for name, (status, emotion, date) in attendance.items():
    ws.append([name, status, emotion, date])

wb.save(excel_path)
print("✅ Attendance updated in Excel:", excel_path)




